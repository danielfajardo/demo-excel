import time as t
import argparse
import openpyxl
import os


def copy_test_plan(filename_origin, filename_destination, solution2=True):
    """ Copy test plan from excel A to B

    High level function that copies data from origin to destination excel file.

    Args:
        filename_origin (str): Original test plan where the data will be extracted
        filename_destination (str): Destination test plan where the data will be copied
        solution2 (bool): True for using solution2. False for using solution1.

    Returns:
        float: Elapsed execution time

    """
    wb_origin, ws_origin = _open_test_plan(filename_origin)
    wb_destination, ws_destination = _open_test_plan(filename_destination)

    if solution2:
        elapsed_time = _solution2(filename_destination, wb_destination, ws_destination, wb_origin, ws_origin)
    else:
        elapsed_time = _solution1(filename_destination, wb_destination, ws_destination, wb_origin, ws_origin)

    return elapsed_time


def _open_test_plan(filename):
    """ Open test plan

    Function that returns the workbook and the desired worksheet to be copied.

    Args:
        filename (str): Name of the excel file

    Returns:
        wb: Workbook of the filename
        ws: TestPlan worksheet of the filename

    """
    wb = openpyxl.load_workbook(filename)
    ws = wb['TestPlan']

    return wb, ws


def _solution1(filename, wb_destination, ws_destination, wb_origin, ws_origin):
    """ Solution1 - very first solution

    Function that copies data from origin to destination excel file, using the test case name as key.

    Args:
        filename (str): Filename destination
        wb_destination (openpyxl.worsheet): Destination workbook
        ws_destination (openpyxl.worsheet): Destination TestPlan worksheet
        wb_origin (openpyxl.worsheet): Origin workbook
        ws_origin (openpyxl.worsheet): Destination TestPlan worksheet

    Returns:
        float: Elapsed execution time

    """
    t_init = t.time()

    for row in range(17, ws_destination.max_row + 1):
        current_cell = ws_destination.cell(row=row, column=1).value
        print(current_cell)

        if current_cell not in [None, 'Test Case ID'] and current_cell.find(".TS.") == -1:
            not_found = True
            pos = 17
            while not_found and pos <= ws_origin.max_row:
                if current_cell == ws_origin.cell(row=pos, column=1).value:
                    for delta in range(0, 7):
                        ws_destination.cell(row=row, column=11 + delta).value = ws_origin.cell(row=pos,
                                                                                               column=11 + delta).value
                    not_found = False
                pos += 1

    wb_destination.save(filename)

    return t.time() - t_init


def _solution2(filename, wb_destination, ws_destination, wb_origin, ws_origin):
    """ Solution2 - using generators

    Function that copies data from origin to destination excel file, using the test case name as key.
    This is more efficient than solution1.

    Args:
        filename (str): Filename destination
        wb_destination (openpyxl.worsheet): Destination workbook
        ws_destination (openpyxl.worsheet): Destination TestPlan worksheet
        wb_origin (openpyxl.worsheet): Origin workbook
        ws_origin (openpyxl.worsheet): Destination TestPlan worksheet

    Returns:
        float: Elapsed execution time

    """
    t_init = t.time()

    for row in ws_destination.iter_rows(min_row=17, max_row=ws_destination.max_row, min_col=1, max_col=17):
        test_name = row[0].value

        if test_name not in [None, 'Test Case ID'] and test_name.find(".TS.") == -1:
            print(test_name)
            for entry in ws_origin.iter_rows(min_row=17, max_row=ws_origin.max_row, min_col=1, max_col=17):
                if test_name == entry[0].value:
                    for delta in range(0, 7):
                        row[10 + delta].value = entry[10 + delta].value
                    break

    wb_destination.save(filename)

    return t.time() - t_init


if __name__ == '__main__':

    parser = argparse.ArgumentParser(prog='Excel demo',
                                     description='This code copies matching values from one excel to another one',
                                     add_help=True)
    parser.add_argument('-fo', metavar='filename origin', help='Name of the origin excel file', required=True)
    parser.add_argument('-fd', metavar='filename destination', help='Name of the destination excel file', required=True)
    parser.add_argument('-s1', dest='solution2', action='store_false')
    parser.add_argument('-s2', dest='solution2', action='store_true')
    parser.set_defaults(solution2=True)

    args = parser.parse_args()

    script_dir_fd = ""
    script_dir_fo = ""

    if not os.path.isfile(args.fo):
        # Assuming that are relative to the script dir
        script_dir_fo = os.path.dirname(os.path.realpath(__file__))

    if not os.path.isfile(args.fd):
        # Assuming that are relative to the script dir
        script_dir_fd = os.path.dirname(os.path.realpath(__file__))

    fo = os.path.join(script_dir_fo, args.fo)
    fd = os.path.join(script_dir_fd, args.fd)

    try:
        elapsed_time = copy_test_plan(fo, fd, args.solution2)
    except FileNotFoundError as msg:
        print(f"File not found! \nReason: {msg}")
    else:
        print(f"Elapsed time {elapsed_time}")
